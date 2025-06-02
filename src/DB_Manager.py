# 이 파일은 리팩토링되어 실제 코드는 app/manager.py, app/schema.py, app/loading.py 등에서 확인하세요.
# 프로그램 실행은 main.py를 사용하세요.

# 작성자: Levi Beak
# 최종 수정일: 2025-05-01

import pandas as pd
import glob
import os
import sqlite3
import tkinter as tk
from tkinter import filedialog, ttk, messagebox, simpledialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from datetime import datetime
import sys
import openpyxl  # Excel 보고서 생성을 위한 모듈
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# 로컬 모듈
from db_schema import DBSchema
from default_db_helpers import add_default_db_functions_to_class
from change_history_helpers import add_change_history_functions_to_class
from common_utils import verify_password, change_maintenance_password
from LoadingDialog import LoadingDialog

class DBManager:
    def __init__(self):
        # 1. 상태 변수 및 기본 속성 초기화
        self.maint_mode = False
        self.selected_equipment_type_id = None
        self.file_names = []
        self.folder_path = ""
        self.merged_df = None
        self.context_menu = None

        # 2. DB 스키마 및 기능 확장 초기화 (예외 발생 시 메시지 출력)
        try:
            self.db_schema = DBSchema()
        except Exception as e:
            print(f"DB 스키마 초기화 실패: {str(e)}")
            self.db_schema = None
        add_default_db_functions_to_class(DBManager)
        add_change_history_functions_to_class(DBManager)

        # 3. 메인 윈도우 및 UI 요소 초기화
        self.window = tk.Tk()
        self.window.title("DB Manager")
        self.window.geometry("1300x800")
        try:
            application_path = sys._MEIPASS if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
            icon_path = os.path.join(application_path, "resources", "icons", "db_compare.ico")
            self.window.iconbitmap(icon_path)
        except Exception as e:
            print(f"아이콘 로드 실패: {str(e)}")

        # 메뉴바 생성
        self.create_menu()

        # 상태바
        self.status_bar = ttk.Label(self.window, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # 메인 노트북 및 비교 노트북
        self.main_notebook = ttk.Notebook(self.window)
        self.main_notebook.pack(expand=True, fill=tk.BOTH)
        self.comparison_notebook = ttk.Notebook(self.main_notebook)
        self.main_notebook.add(self.comparison_notebook, text="DB 비교")

        # 로그 표시 영역 및 스크롤바
        self.log_text = tk.Text(self.window, height=5, state=tk.DISABLED)
        self.log_text.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=5)
        log_scrollbar = ttk.Scrollbar(self.log_text, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scrollbar.set)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 단축키 바인딩 (중복 제거, 함수 직접 참조)
        self.window.bind('<F1>', self.show_user_guide)
        for key in ('<Control-o>', '<Control-O>'):
            self.window.bind(key, self.load_folder)

        # 비교 탭 생성 및 상태 초기화
        self.create_comparison_tabs()
        self.status_bar.config(text="Ready")
        self.update_log("DB Manager 초기화 완료")
        if self.db_schema:
            self.update_log("로컬 데이터베이스 초기화 완료")
            self.update_log("Default DB 관리 기능 준비 완료")
        else:
            self.update_log("DB 스키마 초기화 실패")

    def update_log(self, message):
        """로그 표시 영역에 메시지를 추가합니다."""
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}\n")
        self.log_text.see(tk.END)  # 스크롤을 마지막으로 이동
        self.log_text.configure(state=tk.DISABLED)
    
    def create_menu(self):
        """메뉴바를 생성합니다."""
        menubar = tk.Menu(self.window)
        
        # 파일 메뉴
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="폴더 열기 (Ctrl+O)", command=self.load_folder)
        file_menu.add_separator()
        file_menu.add_command(label="보고서 내보내기", command=self.export_report)
        file_menu.add_separator()
        file_menu.add_command(label="종료", command=self.window.quit)
        menubar.add_cascade(label="파일", menu=file_menu)
        
        # 도구 메뉴
        tools_menu = tk.Menu(menubar, tearoff=0)
        tools_menu.add_command(label="Maintenance Mode", command=self.toggle_maint_mode)
        tools_menu.add_command(label="비밀번호 변경", command=self.show_change_password_dialog)
        menubar.add_cascade(label="도구", menu=tools_menu)
        
        # 도움말 메뉴
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="사용 설명서 (F1)", command=self.show_user_guide)
        help_menu.add_separator()
        help_menu.add_command(label="프로그램 정보", command=self.show_about)
        menubar.add_cascade(label="도움말", menu=help_menu)
        
        self.window.config(menu=menubar)

    def toggle_maint_mode(self):
        """유지보수 모드 토글"""
        if self.maint_mode:
            # 유지보수 모드 비활성화
            self.maint_mode = False
            self.status_bar.config(text="유지보수 모드 비활성화")
            self.update_log("유지보수 모드가 비활성화되었습니다.")
            # 유지보수 관련 기능 비활성화
            self.disable_maint_features()
        else:
            # 유지보수 모드 활성화 (비밀번호 확인)
            password = simpledialog.askstring("유지보수 모드", "비밀번호를 입력하세요:", show="*")
            if password is None:  # 취소 버튼 클릭 시
                return
            
            if verify_password(password):
                # 로딩 다이얼로그 표시
                loading_dialog = LoadingDialog(self.window)
                loading_dialog.update_progress(10, "유지보수 모드 활성화 중...")
                
                self.maint_mode = True
                self.status_bar.config(text="유지보수 모드 활성화")
                self.update_log("유지보수 모드가 활성화되었습니다.")
                
                # 유지보수 관련 기능 활성화
                try:
                    loading_dialog.update_progress(30, "기본 DB 관리 기능 초기화 중...")
                    self.enable_maint_features()
                    loading_dialog.update_progress(100, "완료")
                    loading_dialog.close()
                except Exception as e:
                    loading_dialog.close()
                    messagebox.showerror("오류", f"유지보수 모드 활성화 중 오류 발생: {str(e)}")
            else:
                messagebox.showerror("오류", "비밀번호가 일치하지 않습니다.")
                return
        
        # Default DB 관련 UI 요소 상태 업데이트
        self.update_default_db_ui_state()
    
    def show_change_password_dialog(self):
        """유지보수 모드 비밀번호 변경 다이얼로그를 표시합니다."""
        # 현재 비밀번호 확인
        current_password = simpledialog.askstring("비밀번호 변경", "현재 비밀번호를 입력하세요:", show="*")
        if current_password is None:  # 취소 버튼 클릭 시
            return
            
        if not verify_password(current_password):
            messagebox.showerror("오류", "현재 비밀번호가 일치하지 않습니다.")
            return
            
        # 새 비밀번호 입력
        new_password = simpledialog.askstring("비밀번호 변경", "새 비밀번호를 입력하세요:", show="*")
        if new_password is None:  # 취소 버튼 클릭 시
            return
            
        # 새 비밀번호 확인
        confirm_password = simpledialog.askstring("비밀번호 변경", "새 비밀번호를 다시 입력하세요:", show="*")
        if confirm_password is None:  # 취소 버튼 클릭 시
            return
            
        if new_password != confirm_password:
            messagebox.showerror("오류", "새 비밀번호가 일치하지 않습니다.")
            return
            
        # 비밀번호 변경
        if change_maintenance_password(current_password, new_password):
            messagebox.showinfo("성공", "비밀번호가 성공적으로 변경되었습니다.")
            self.update_log("유지보수 모드 비밀번호가 변경되었습니다.")
        else:
            messagebox.showerror("오류", "비밀번호 변경에 실패했습니다.")
    
    def update_default_db_ui_state(self):
        """유지보수 모드에 따라 Default DB 관련 UI 요소들의 상태를 업데이트합니다."""
        # Default DB 후보 표시 체크박스 상태 업데이트
        if hasattr(self, 'show_default_candidates_cb'):
            if self.maint_mode:
                self.show_default_candidates_cb.configure(state="normal")
            else:
                self.show_default_candidates_var.set(False)  # 체크 해제
                self.show_default_candidates_cb.configure(state="disabled")
                self.update_comparison_view()  # 뷰 업데이트
        
        # 컨텍스트 메뉴 상태 업데이트
        self.update_comparison_context_menu_state()
        
        # 비교 탭 UI 갱신 (유지보수 모드 변경 시)
        self.update_all_tabs()

    def enable_maint_features(self):
        """유지보수 모드 활성화 시 필요한 기능을 활성화합니다. (최적화: 모든 무거운 작업을 스레드로 분리)"""
        import threading
        loading_dialog = LoadingDialog(self.window)

        def worker():
            try:
                # 1. UI 관련 활성화 (after로 안전하게)
                if hasattr(self, 'notebook') and self.notebook:
                    for tab_id in range(self.notebook.index('end')):
                        if self.notebook.tab(tab_id, 'text') in ["Default DB 관리", "QC 검수", "변경 이력 관리"]:
                            self.window.after(0, lambda tab_id=tab_id: self.notebook.tab(tab_id, state='normal'))
                for widget_name in ['add_equipment_button', 'add_parameter_button', 'edit_button', 'delete_button']:
                    if hasattr(self, widget_name) and getattr(self, widget_name):
                        self.window.after(0, lambda wn=widget_name: getattr(self, wn).config(state='normal'))
                if hasattr(self, 'equipment_tree'):
                    self.window.after(0, lambda: self.equipment_tree.bind('<Double-1>', self.on_tree_double_click))
                    self.window.after(0, lambda: self.equipment_tree.bind('<Button-3>', self.show_context_menu))

                # 2. 무거운 탭 생성 (QC/Default DB/변경이력)
                loading_dialog.update_progress(10, "QC 탭 생성 중...")
                if hasattr(self, "create_qc_check_tab"):
                    self.create_qc_check_tab()
                elif hasattr(self, "create_qc_tab"):
                    self.create_qc_tab()

                loading_dialog.update_progress(40, "Default DB 관리 탭 생성 중...")
                if hasattr(self, "create_default_db_tab"):
                    self.create_default_db_tab()

                loading_dialog.update_progress(70, "변경 이력 관리 탭 생성 중...")
                if hasattr(self, "create_change_history_tab"):
                    self.create_change_history_tab()

                # 3. 완료 처리
                self.window.after(0, lambda: self.update_log("Maintenance Mode가 활성화되었습니다."))
                loading_dialog.update_progress(100, "완료!")
                self.window.after(0, loading_dialog.close)
            except Exception as e:
                self.window.after(0, loading_dialog.close)
                self.window.after(0, lambda e=e: messagebox.showerror("오류", f"유지보수 모드 활성화 중 오류 발생: {str(e)}"))

        threading.Thread(target=worker, daemon=True).start()

        
    def disable_maint_features(self):
        """Maintenance Mode 비활성화 시 관련 기능을 제거합니다."""
        # 유지보수 관련 탭 제거
        for tab in self.main_notebook.tabs():
            tab_text = self.main_notebook.tab(tab, "text")
            if tab_text == "QC 검수" or tab_text == "Default DB 관리":
                self.main_notebook.forget(tab)
        
        # DB 비교 탭 선택
        for tab in self.main_notebook.tabs():
            if self.main_notebook.tab(tab, "text") == "DB 비교":
                self.main_notebook.select(tab)
        """개선된 QC 검수 탭을 생성합니다. 이 탭은 유지보수 모드에서 활성화됩니다."""
        # QC 검수 탭이 없을 때만 생성
        if not any(self.main_notebook.tab(tab, "text") == "QC 검수" 
                  for tab in self.main_notebook.tabs()):
            qc_tab = ttk.Frame(self.main_notebook)
            self.main_notebook.add(qc_tab, text="QC 검수")
            
            # 메인 프레임 구성
            main_frame = ttk.Frame(qc_tab)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # 상태 표시 프레임
            status_frame = ttk.LabelFrame(main_frame, text="검증 상태")
            status_frame.pack(fill=tk.X, padx=5, pady=5)
            
            # 상태 표시 레이아웃
            status_content = ttk.Frame(status_frame)
            status_content.pack(fill=tk.X, padx=10, pady=10)
            
            # 상태 표시 라벨
            ttk.Label(status_content, text="로드된 파일:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
            self.qc_loaded_files_label = ttk.Label(status_content, text="없음")
            self.qc_loaded_files_label.grid(row=0, column=1, sticky="w", padx=5, pady=2)
            
            ttk.Label(status_content, text="선택된 장비 유형:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
            self.qc_equipment_type_label = ttk.Label(status_content, text="없음")
            self.qc_equipment_type_label.grid(row=1, column=1, sticky="w", padx=5, pady=2)
            
            ttk.Label(status_content, text="검증 상태:").grid(row=2, column=0, sticky="w", padx=5, pady=2)
            self.qc_validation_status_label = ttk.Label(status_content, text="검증 대기 중")
            self.qc_validation_status_label.grid(row=2, column=1, sticky="w", padx=5, pady=2)
            
            # 장비 유형 선택 프레임
            equipment_frame = ttk.LabelFrame(main_frame, text="장비 유형 선택")
            equipment_frame.pack(fill=tk.X, padx=5, pady=5)
            
            equipment_content = ttk.Frame(equipment_frame)
            equipment_content.pack(fill=tk.X, padx=10, pady=10)
            
            # 장비 유형 선택 콤보박스
            ttk.Label(equipment_content, text="장비 유형:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
            self.qc_equipment_type_combo = ttk.Combobox(equipment_content, state="readonly")
            self.qc_equipment_type_combo.grid(row=0, column=1, sticky="ew", padx=5, pady=2)
            self.qc_equipment_type_combo.bind("<<ComboboxSelected>>", self.on_qc_equipment_type_selected)
            
            # 검증 시작 버튼
            ttk.Button(equipment_content, text="검증 시작", command=self.start_validation).grid(row=0, column=2, padx=10, pady=2)
            
            # 추가 옵션
            options_frame = ttk.Frame(equipment_content)
            options_frame.grid(row=1, column=0, columnspan=3, sticky="w", padx=5, pady=5)
            
            self.show_all_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(options_frame, text="모든 항목 표시", variable=self.show_all_var,
                           command=self.refresh_validation_view).pack(side=tk.LEFT, padx=5)
            
            self.highlight_errors_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(options_frame, text="오류 항목 강조", variable=self.highlight_errors_var,
                           command=self.refresh_validation_view).pack(side=tk.LEFT, padx=5)
            
            # 검증 결과 프레임
            results_frame = ttk.LabelFrame(main_frame, text="검증 결과")
            results_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # 검증 결과 테이블
            table_frame = ttk.Frame(results_frame)
            table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # 테이블 객체 (트리뷰) 생성
            columns = ("parameter", "model", "value", "default", "min", "max", "status")
            self.qc_result_tree = ttk.Treeview(table_frame, columns=columns, show="headings")
            
            # 헤더 설정
            headers = {
                "parameter": "DB_ItemName",
                "model": "모델", 
                "value": "값",
                "default": "기본값",
                "min": "최소값",
                "max": "최대값",
                "status": "상태"
            }
            
            for col, header in headers.items():
                self.qc_result_tree.heading(col, text=header)
                # 적절한 너비 설정
                width = 70
                if col in ["parameter", "model"]:
                    width = 100
                elif col == "value":
                    width = 80
                self.qc_result_tree.column(col, width=width)
            
            # 스크롤바 추가
            tree_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.qc_result_tree.yview)
            self.qc_result_tree.configure(yscrollcommand=tree_scrollbar.set)
            
            # 테이블 및 스크롤바 배치
            self.qc_result_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 버튼 프레임
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill=tk.X, padx=5, pady=10)
            
            ttk.Button(button_frame, text="보고서 내보내기", command=self.export_validation_report).pack(side=tk.RIGHT, padx=5)
            
            # 기본 장비 유형 목록 로드
            self.load_equipment_types_for_qc()
        # Placeholder for exporting the QC validation report
        print("Exporting QC Validation Report")
        # TODO: Implement the actual report export logic here
        pass

    def create_default_db_tab(self):
        """유지보수 모드에서 보여질 Default DB 관리 탭을 생성합니다."""
        # 탭이 이미 존재하는지 확인
        for tab in self.main_notebook.tabs():
            if self.main_notebook.tab(tab, "text") == "Default DB 관리":
                self.main_notebook.select(tab)  # 탭이 존재하면 해당 탭 선택
                return
                
        # 새 탭 생성
        default_db_frame = ttk.Frame(self.main_notebook)
        self.main_notebook.add(default_db_frame, text="Default DB 관리")
        self.main_notebook.select(default_db_frame)
        
        # 레이아웃 구성: 좌측은 장비 유형 목록, 우측은 DB 값 관리
        left_frame = ttk.Frame(default_db_frame, width=300)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        
        right_frame = ttk.Frame(default_db_frame)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 좌측 프레임: 장비 유형 관리
        ttk.Label(left_frame, text="장비 유형 관리", font=("Helvetica", 12, "bold")).pack(pady=10)
        
        # 장비 유형 목록 
        self.equipment_type_frame = ttk.Frame(left_frame)
        self.equipment_type_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.equipment_type_listbox = tk.Listbox(self.equipment_type_frame, width=30, height=15)
        equipment_scrollbar = ttk.Scrollbar(self.equipment_type_frame, orient="vertical", command=self.equipment_type_listbox.yview)
        self.equipment_type_listbox.config(yscrollcommand=equipment_scrollbar.set)
        
        self.equipment_type_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        equipment_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.equipment_type_listbox.bind('<<ListboxSelect>>', self.on_equipment_type_select)
        
        # 장비 유형 관리 버튼 프레임
        type_btn_frame = ttk.Frame(left_frame)
        type_btn_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(type_btn_frame, text="추가", command=self.add_equipment_type).pack(side=tk.LEFT, padx=2)
        ttk.Button(type_btn_frame, text="수정", command=self.edit_equipment_type).pack(side=tk.LEFT, padx=2)
        ttk.Button(type_btn_frame, text="삭제", command=self.delete_equipment_type).pack(side=tk.LEFT, padx=2)
        
        # 우측 프레임: DB 값 관리
        ttk.Label(right_frame, text="Default DB 값 관리", font=("Helvetica", 12, "bold")).pack(pady=10)
        
        # DB 값 테이블 (Treeview)
        db_values_frame = ttk.Frame(right_frame)
        db_values_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        columns = ("parameter", "default", "min", "max")
        self.db_values_tree = ttk.Treeview(db_values_frame, columns=columns, show="headings", height=15)
        
        # 헤더 설정
        self.db_values_tree.heading("parameter", text="DB_ItemName")
        self.db_values_tree.heading("default", text="기본값")
        self.db_values_tree.heading("min", text="최소값")
        self.db_values_tree.heading("max", text="최대값")
        
        # 열 너비 설정
        self.db_values_tree.column("parameter", width=150)
        self.db_values_tree.column("default", width=100)
        self.db_values_tree.column("min", width=100)
        self.db_values_tree.column("max", width=100)
        
        # 스크롤바 추가
        db_values_scrollbar = ttk.Scrollbar(db_values_frame, orient="vertical", command=self.db_values_tree.yview)
        self.db_values_tree.configure(yscrollcommand=db_values_scrollbar.set)
        
        self.db_values_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        db_values_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # DB 값 관리 버튼 프레임
        values_btn_frame = ttk.Frame(right_frame)
        values_btn_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(values_btn_frame, text="추가", command=self.add_default_value).pack(side=tk.LEFT, padx=2)
        ttk.Button(values_btn_frame, text="수정", command=self.edit_default_value).pack(side=tk.LEFT, padx=2)
        ttk.Button(values_btn_frame, text="삭제", command=self.delete_default_value).pack(side=tk.LEFT, padx=2)
        ttk.Button(values_btn_frame, text="임포트", command=self.import_default_values).pack(side=tk.LEFT, padx=20)
        ttk.Button(values_btn_frame, text="익스포트", command=self.export_default_values).pack(side=tk.LEFT, padx=2)
        
        # 장비 유형 목록 초기화
        self.update_equipment_type_list()

    def load_equipment_types_for_qc(self):
        # Placeholder for loading equipment types into the QC tab combobox
        print("Loading equipment types for QC tab")
        try:
            types = self.db_schema.get_equipment_types()
            type_names = [t[1] for t in types] # Assuming type name is the second element
            self.qc_equipment_type_combo['values'] = type_names
            if type_names:
                self.qc_equipment_type_combo.current(0)
            else:
                self.qc_equipment_type_combo.set('')
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load equipment types: {e}")
            self.qc_equipment_type_combo['values'] = []
            self.qc_equipment_type_combo.set('')

    def create_comparison_tabs(self):
        """비교 관련 탭 생성"""
        # 격자 뷰 탭 생성
        self.create_grid_view_tab()
        # 다른 값만 보기 탭 생성
        self.create_diff_only_tab()
        # 보고서 탭 생성
        self.create_report_tab()

    def load_folder(self):
        # 파일 확장자 필터 설정
        filetypes = [
            ("DB 파일", "*.txt;*.db;*.csv"),
            ("텍스트 파일", "*.txt"),
            ("CSV 파일", "*.csv"),
            ("DB 파일", "*.db"),
            ("모든 파일", "*.*")
        ]
        
        # 파일 선택 다이얼로그
        files = filedialog.askopenfilenames(
            title="📂 DB 파일을 선택하세요",
            filetypes=filetypes,
            initialdir=self.folder_path if self.folder_path else None
        )
        
        if not files:
            self.status_bar.config(text="파일 선택이 취소되었습니다.")
            return

        # 로딩 다이얼로그 생성
        loading_dialog = LoadingDialog(self.window)
        
        try:
            df_list = []
            self.file_names = []
            total_files = len(files)
            
            # 초기 진행률 표시
            loading_dialog.update_progress(0, "파일 로딩 준비 중...")
            
            for idx, file in enumerate(files, 1):
                try:
                    # 파일 로딩 진행률 (0-70%)
                    progress = (idx / total_files) * 70
                    loading_dialog.update_progress(
                        progress,
                        f"파일 로딩 중... ({idx}/{total_files})"
                    )
                    
                    file_name = os.path.basename(file)
                    base_name = os.path.splitext(file_name)[0]
                    ext = os.path.splitext(file_name)[1].lower()
                    
                    # 파일 형식에 따른 로드
                    if ext == '.txt':
                        df = pd.read_csv(file, delimiter="\t", dtype=str)
                    elif ext == '.csv':
                        df = pd.read_csv(file, dtype=str)
                    elif ext == '.db':
                        conn = sqlite3.connect(file)
                        df = pd.read_sql("SELECT * FROM main_table", conn)
                        conn.close()
                    
                    df["Model"] = base_name
                    df_list.append(df)
                    self.file_names.append(base_name)
                    
                except Exception as e:
                    messagebox.showwarning(
                        "경고", 
                        f"'{file_name}' 파일 로드 중 오류 발생:\n{str(e)}"
                    )
            
            if df_list:
                # 선택된 파일들의 폴더 경로 저장
                self.folder_path = os.path.dirname(files[0])
                
                # 데이터프레임 병합
                loading_dialog.update_progress(75, "데이터 병합 중...")
                self.merged_df = pd.concat(df_list, ignore_index=True)
                
                # UI 업데이트
                loading_dialog.update_progress(85, "화면 업데이트 중...")
                self.update_all_tabs()
                
                loading_dialog.update_progress(100, "완료!")
                
                # 완료 메시지
                loading_dialog.close()
                messagebox.showinfo(
                    "로드 완료",
                    f"총 {len(df_list)}개의 DB 파일을 성공적으로 로드했습니다.\n"
                    f"• 폴더: {self.folder_path}\n"
                    f"• 파일: {', '.join(self.file_names)}"
                )
                
                # 상태 표시줄 업데이트
                self.status_bar.config(
                    text=f"총 {len(df_list)}개의 DB 파일이 로드되었습니다. "
                         f"(폴더: {os.path.basename(self.folder_path)})"
                )
            else:
                loading_dialog.close()
                messagebox.showerror("오류", "파일을 로드할 수 없습니다.")
                self.status_bar.config(text="파일 로드 실패")
                
        except Exception as e:
            loading_dialog.close()
            messagebox.showerror("오류", f"예기치 않은 오류가 발생했습니다:\n{str(e)}")

    def update_all_tabs(self):
        # 기존 탭 제거
        for tab in self.comparison_notebook.winfo_children():
            tab.destroy()
        
        # 탭 다시 생성
        self.create_comparison_tabs()

    def create_grid_view_tab(self):
        grid_frame = ttk.Frame(self.comparison_notebook)
        self.comparison_notebook.add(grid_frame, text="📑 격자 뷰")
        
        grid_frame.rowconfigure([0, 1, 2], weight=1)
        grid_frame.columnconfigure([0, 1], weight=1)
        
        def count_items(df, module=None, part=None):
            if module is not None and part is not None:
                return df[(df["Module"] == module) & (df["Part"] == part)]["ItemName"].nunique()
            elif module is not None:
                return df[df["Module"] == module]["ItemName"].nunique()
            else:
                return df["ItemName"].nunique()
        
        max_columns = 2
        for idx, file_name in enumerate(self.file_names[:6]):
            row = idx // max_columns
            col = idx % max_columns
            
            frame = ttk.LabelFrame(grid_frame, text=f"📌 {file_name}")
            frame.grid(row=row, column=col, padx=5, pady=5, sticky="nsew")
            frame.rowconfigure(0, weight=1)
            frame.columnconfigure(0, weight=1)
            
            container = ttk.Frame(frame)
            container.pack(fill=tk.BOTH, expand=True)
            
            # 트리뷰 추가
            tree = ttk.Treeview(container)
            tree["columns"] = ("Module", "Part", "ItemName", "ItemValue")
            tree.heading("#0", text="구조", anchor="w")
            tree.heading("Module", text="Module", anchor="w")
            tree.heading("Part", text="Part", anchor="w")
            tree.heading("ItemName", text="Item Name", anchor="w")
            tree.heading("ItemValue", text="Item Value", anchor="w")
            
            # 스크롤바 추가
            v_scroll = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
            h_scroll = ttk.Scrollbar(container, orient="horizontal", command=tree.xview)
            tree.configure(yscroll=v_scroll.set, xscroll=h_scroll.set)
            
            v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
            h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
            tree.pack(expand=True, fill=tk.BOTH)
            
            # 데이터 추가
            if self.merged_df is not None:
                model_data = self.merged_df[self.merged_df["Model"] == file_name]
                for module in model_data["Module"].unique():
                    module_count = count_items(model_data, module=module)
                    module_node = tree.insert("", "end", text=f"📁 {module} ({module_count})", open=False)
                    
                    module_data = model_data[model_data["Module"] == module]
                    for part in module_data["Part"].unique():
                        part_count = count_items(module_data, module=module, part=part)
                        part_node = tree.insert(module_node, "end", text=f"📂 {part} ({part_count})", open=False)
                        
                        part_data = module_data[module_data["Part"] == part]
                        for _, row in part_data.iterrows():
                            item_name = row["ItemName"]
                            item_value = row["ItemValue"]
                            tree.insert(part_node, "end", text=item_name, 
                                      values=(module, part, item_name, item_value))

    # create_comparison_tab 완전 삭제 (DB 값 비교 탭 및 기능 제거)

    def create_diff_only_tab(self):
        import pandas as pd
        import numpy as np
        from tkinter import filedialog

        def treeview_to_dataframe(tree, columns):
            data = [tree.item(item)["values"] for item in tree.get_children()]
            return pd.DataFrame(data, columns=columns)

        def export_treeview_to_excel(tree, columns, default_name):
            df = treeview_to_dataframe(tree, columns)
            file_path = filedialog.asksaveasfilename(
                title="엑셀로 내보내기",
                defaultextension=".xlsx",
                filetypes=[("Excel 파일", ".xlsx"), ("모든 파일", "*.*")],
                initialfile=default_name
            )
            if not file_path:
                return
            df.to_excel(file_path, index=False)
            messagebox.showinfo("완료", f"엑셀 파일이 저장되었습니다.\n{file_path}")
            if messagebox.askyesno("파일 열기", "저장된 파일을 바로 열까요?"):
                import os
                os.startfile(file_path)

        # --- UI 프레임 ---
        diff_frame = ttk.Frame(self.comparison_notebook)
        self.comparison_notebook.add(diff_frame, text="🔍 차이점만 보기")

        control_frame = ttk.Frame(diff_frame)
        control_frame.pack(fill=tk.X, padx=5, pady=5)

        # 검색 입력란
        search_var = tk.StringVar()
        ttk.Label(control_frame, text="검색: ").pack(side=tk.LEFT)
        search_entry = ttk.Entry(control_frame, textvariable=search_var, width=20)
        search_entry.pack(side=tk.LEFT, padx=2)

        # 엑셀로 내보내기 버튼
        export_btn = ttk.Button(control_frame, text="엑셀로 내보내기",
            command=lambda: export_treeview_to_excel(tree, columns, "DB_DiffOnly.xlsx"))
        export_btn.pack(side=tk.RIGHT, padx=5)

        # Treeview 컬럼 설정
        columns = ["Module", "Part", "ItemName"] + self.file_names
        tree = ttk.Treeview(diff_frame, columns=columns, show="headings", height=30)
        for col in columns:
            tree.heading(col, text=col,
                command=lambda c=col: sort_treeview(tree, c, False))
            tree.column(col, width=120 if col in ["Module", "Part", "ItemName"] else 150, anchor="w")
        tree.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        # 스크롤바
        v_scroll = ttk.Scrollbar(diff_frame, orient="vertical", command=tree.yview)
        h_scroll = ttk.Scrollbar(diff_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscroll=v_scroll.set, xscroll=h_scroll.set)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)

        # 데이터 렌더링 (최적화)
        def render():
            tree.delete(*tree.get_children())
            if self.merged_df is not None:
                df = self.merged_df.copy()
                keyword = search_var.get().strip()
                if keyword:
                    mask = df.apply(lambda row: keyword in str(row.values), axis=1)
                    df = df[mask]
                grouped = df.groupby(["Module", "Part", "ItemName"])
                for (module, part, item_name), group in grouped:
                    values = [module, part, item_name]
                    model_values = []
                    for model in self.file_names:
                        model_value = group[group["Model"] == model]["ItemValue"].values
                        value = model_value[0] if len(model_value) > 0 else "-"
                        values.append(value)
                        model_values.append(value)
                    if len(set(model_values)) > 1:
                        tree.insert("", "end", values=values, tags=("different",))
                tree.tag_configure("different", background="light yellow")
        
        def sort_treeview(tree, col, reverse):
            l = [(tree.set(k, col), k) for k in tree.get_children("")]
            try:
                l.sort(key=lambda t: float(t[0]), reverse=reverse)
            except ValueError:
                l.sort(reverse=reverse)
            for index, (val, k) in enumerate(l):
                tree.move(k, '', index)
            tree.heading(col, command=lambda: sort_treeview(tree, col, not reverse))

        search_var.trace_add('write', lambda *args: render())
        render()

        comparison_frame = ttk.Frame(self.comparison_notebook)
        self.comparison_notebook.add(comparison_frame, text="DB 값 비교")
        
        # 트리뷰 행간격 스타일 조정 (글자 크기에 맞게)
        style = ttk.Style()
        style.configure("Custom.Treeview", rowheight=22)
        
        # 상단 컨트롤 프레임 추가
        control_frame = ttk.Frame(comparison_frame)
        control_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 유지보수 모드일 때만 체크박스, 버튼 생성 (Default DB 추가 옵션 프레임은 제거)
        if self.maint_mode:
            # 모두 선택 체크박스
            self.select_all_var = tk.BooleanVar(value=False)
            self.select_all_cb = ttk.Checkbutton(
                control_frame,
                text="모두 선택",
                variable=self.select_all_var,
                command=self.toggle_select_all_checkboxes
            )
            self.select_all_cb.pack(side=tk.LEFT, padx=5)
        
        # 상태 표시 레이블: 유지보수 모드면 '선택된 항목', 아니면 '값이 다른 항목'
        if self.maint_mode:
            self.selected_count_label = ttk.Label(control_frame, text="선택된 항목: 0개")
            self.selected_count_label.pack(side=tk.RIGHT, padx=10)
            # 유지보수 모드일 때만 'Default DB로 전송' 버튼 생성
            self.send_to_default_btn = ttk.Button(
                control_frame,
                text="Default DB로 전송",
                command=self.add_to_default_db
            )
            self.send_to_default_btn.pack(side=tk.RIGHT, padx=10)
        else:
            pass
        
    def export_treeview_to_excel(tree, columns, default_name):
        df = treeview_to_dataframe(tree, columns)
        file_path = filedialog.asksaveasfilename(
            title="엑셀로 내보내기",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", ".xlsx"), ("모든 파일", "*.*")],
            initialfile=default_name
        )
        if not file_path:
            return
        df.to_excel(file_path, index=False)
        messagebox.showinfo("완료", f"엑셀 파일이 저장되었습니다.\n{file_path}")
        if messagebox.askyesno("파일 열기", "저장된 파일을 바로 열까요?"):
            import os
            os.startfile(file_path)
        check = self.select_all_var.get()
        for item in self.comparison_tree.get_children():
            values = list(self.comparison_tree.item(item, "values"))
            if len(values) > 0:
                values[0] = "☑" if check else "☐"
                self.comparison_tree.item(item, values=values)
                module, part, item_name = values[1], values[2], values[3]
                item_key = f"{module}_{part}_{item_name}"
                self.item_checkboxes[item_key] = check
        self.update_checked_count()

    def update_comparison_view(self):
        """비교 탭의 데이터를 업데이트합니다. 유지보수 모드에 따라 체크박스 기능을 활성화/비활성화합니다."""
        for item in self.comparison_tree.get_children():
            self.comparison_tree.delete(item)
        saved_checkboxes = self.item_checkboxes.copy()
        self.item_checkboxes.clear()
        if self.maint_mode:
            self.comparison_tree.bind("<ButtonRelease-1>", self.toggle_checkbox)
        else:
            self.comparison_tree.unbind("<ButtonRelease-1>")
        diff_count = 0
        if self.merged_df is not None:
            grouped = self.merged_df.groupby(["Module", "Part", "ItemName"])
            for (module, part, item_name), group in grouped:
                values = []
                if self.maint_mode:
                    checkbox_state = "☐"
                    item_key = f"{module}_{part}_{item_name}"
                    if item_key in saved_checkboxes and saved_checkboxes[item_key]:
                        checkbox_state = "☑"
                    self.item_checkboxes[item_key] = (checkbox_state == "☑")
                    values.append(checkbox_state)
                values.extend([module, part, item_name])
                for model in self.file_names:
                    model_value = group[group["Model"] == model]["ItemValue"].values
                    value = model_value[0] if len(model_value) > 0 else "-"
                    values.append(value)
                tags = []
                model_values = values[(4 if self.maint_mode else 3):]
                if len(set(model_values)) > 1:
                    tags.append("different")
                    diff_count += 1
                is_existing = self.check_if_parameter_exists(module, part, item_name)
                if is_existing:
                    tags.append("existing")
                self.comparison_tree.insert("", "end", values=values, tags=tuple(tags))
            self.comparison_tree.tag_configure("different", background="light yellow")
            self.comparison_tree.tag_configure("existing", foreground="blue")
            if self.maint_mode:
                self.comparison_tree.bind("<ButtonRelease-1>", self.toggle_checkbox)
            self.update_selected_count(None)
        # 유지보수 모드가 아닐 때는 diff_count_label에 값이 다른 항목 개수 표시
        if not self.maint_mode and hasattr(self, 'diff_count_label'):
            self.diff_count_label.config(text=f"값이 다른 항목: {diff_count}개")

    def create_comparison_context_menu(self):
        """DB 비교 탭에서 사용할 컨텍스트 메뉴를 생성합니다."""
        self.comparison_context_menu = tk.Menu(self.window, tearoff=0)
        self.comparison_context_menu.add_command(label="선택한 항목을 Default DB에 추가", command=self.add_to_default_db)
        
        # 오른쪽 클릭 이벤트 바인딩
        self.comparison_tree.bind("<Button-3>", self.show_comparison_context_menu)
        
        # 유지보수 모드가 아닐 때는 메뉴 비활성화
        self.update_comparison_context_menu_state()

    def show_comparison_context_menu(self, event):
        """컨텍스트 메뉴를 표시합니다."""
        # 유지보수 모드가 아닐 때는 메뉴를 표시하지 않음
        if not self.maint_mode:
            return
            
        # 선택된 항목이 있는지 확인
        if not self.comparison_tree.selection():
            return
        
        # 메뉴 표시
        try:
            self.comparison_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.comparison_context_menu.grab_release()
            
    def update_comparison_context_menu_state(self):
        """유지보수 모드에 따라 컨텍스트 메뉴 상태를 업데이트합니다."""
        if hasattr(self, 'comparison_context_menu'):
            state = "normal" if self.maint_mode else "disabled"
            try:
                # 메뉴 아이템 비활성화/활성화
                self.comparison_context_menu.entryconfig("선택한 항목을 Default DB에 추가", state=state)
            except Exception as e:
                self.update_log(f"컨텍스트 메뉴 상태 업데이트 중 오류 발생: {str(e)}")
                
    def toggle_checkbox(self, event):
        """체크박스를 토글합니다. 유지보수 모드가 비활성화된 경우 체크박스 토글을 비활성화합니다."""
        # 유지보수 모드가 아닌 경우 체크박스 토글 비활성화
        if not self.maint_mode:
            return
            
        # 클릭한 항목과 컬럼 정보 가져오기
        region = self.comparison_tree.identify_region(event.x, event.y)
        if region != "cell":
            return
            
        column = self.comparison_tree.identify_column(event.x)
        if column != "#1":  # 체크박스 컬럼이 아니면 무시
            return
            
        item = self.comparison_tree.identify_row(event.y)
        if not item:
            return
            
        # 현재 항목의 값 가져오기
        values = self.comparison_tree.item(item, "values")
        if not values or len(values) < 4:
            return
            
        # 체크박스 상태 토글
        current_state = values[0]
        module, part, item_name = values[1], values[2], values[3]
        item_key = f"{module}_{part}_{item_name}"
        
        # 체크박스 상태 변경
        new_state = "☑" if current_state == "☐" else "☐"
        self.item_checkboxes[item_key] = (new_state == "☑")
        
        # 트리뷰 항목 업데이트
        new_values = list(values)
        new_values[0] = new_state
        self.comparison_tree.item(item, values=new_values)
        
        # 체크된 항목 수 업데이트
        self.update_checked_count()

    def update_selected_count(self, event):
        """선택된 항목 개수를 업데이트합니다."""
        if not self.maint_mode:
            return
        # 체크된 항목 개수 표시
        checked_count = sum(1 for checked in self.item_checkboxes.values() if checked)
        # 체크된 항목이 있으면 체크된 항목 개수 표시, 없으면 선택된 항목 개수 표시
        if checked_count > 0:
            self.selected_count_label.config(text=f"체크된 항목: {checked_count}개")
        else:
            selected_items = self.comparison_tree.selection()
            count = len(selected_items)
            self.selected_count_label.config(text=f"선택된 항목: {count}개")

    def update_checked_count(self):
        """체크된 항목 개수를 업데이트합니다."""
        if not self.maint_mode:
            return
        checked_count = sum(1 for checked in self.item_checkboxes.values() if checked)
        self.selected_count_label.config(text=f"체크된 항목: {checked_count}개")

    def check_if_parameter_exists(self, module, part, item_name):
        """DB_ItemName이 이미 Default DB에 존재하는지 확인합니다."""
        try:
            # 현재 등록된 장비 유형 목록 가져오기
            equipment_types = self.db_schema.get_equipment_types()
            
            # 각 장비 유형에 대해 확인
            for type_id, type_name, _ in equipment_types:
                # 장비 유형 이름과 모듈 이름이 일치하는지 확인
                if type_name.lower() == module.lower():
                    # 해당 장비 유형의 Default 값 목록 가져오기
                    default_values = self.db_schema.get_default_values(type_id)
                    
                    # DB_ItemName이 일치하는지 확인
                    for _, param_name, _, _, _, _ in default_values:
                        # DB_ItemName 형식: Part_ItemName 또는 ItemName
                        if param_name == f"{part}_{item_name}" or param_name == item_name:
                            return True
            
            return False
        except Exception as e:
            self.update_log(f"DB_ItemName 존재 여부 확인 중 오류: {str(e)}")
            return False
            
    
            if not self.maint_mode:
                result = messagebox.askyesno("유지보수 모드 필요", "Default DB를 수정하려면 유지보수 모드가 필요합니다. 지금 활성화하시겠습니까?")
                if result:
                    self.toggle_maint_mode()
                else:
                    return
    
            equipment_types = self.db_schema.get_equipment_types()
            type_names = [name for _, name, _ in equipment_types]
    
            select_dialog = tk.Toplevel(self.window)
            select_dialog.title("장비 유형 선택")
            select_dialog.geometry("350x450") # Adjusted size slightly
            select_dialog.transient(self.window)
            select_dialog.grab_set()
    
            # --- UI Elements for select_dialog ---
            selected_type_var = tk.StringVar()
            list_frame = ttk.Frame(select_dialog)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            ttk.Label(list_frame, text="기존 장비 유형 선택:").pack(anchor=tk.W)
            type_listbox = tk.Listbox(list_frame, height=10, exportselection=False)
            for name_only in type_names:
                type_listbox.insert(tk.END, name_only)
            if type_names:
                type_listbox.selection_set(0)
            type_listbox.pack(fill=tk.BOTH, expand=True, pady=5)
    
            new_type_frame = ttk.Frame(select_dialog)
            new_type_frame.pack(fill=tk.X, padx=10, pady=5)
            ttk.Label(new_type_frame, text="새 장비 유형 추가:").pack(side=tk.LEFT)
            new_type_entry = ttk.Entry(new_type_frame)
            new_type_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
            
            button_frame = ttk.Frame(select_dialog)
            button_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=10, pady=10)
            # --- End UI Elements ---
    
            def on_confirm():
                nonlocal equipment_types # To allow modification if a new type is added
                chosen_type_name = ""
                
                # Determine selected or new equipment type
                if type_listbox.curselection():
                    chosen_type_name = type_listbox.get(type_listbox.curselection())
                elif new_type_entry.get().strip():
                    chosen_type_name = new_type_entry.get().strip()
                    try:
                        self.db_schema.add_equipment_type(chosen_type_name, "")
                        equipment_types = self.db_schema.get_equipment_types() # Refresh
                    except Exception as e:
                        messagebox.showerror("오류", f"새 장비 유형 추가 중 오류 발생: {str(e)}", parent=select_dialog)
                        return
                else:
                    messagebox.showinfo("알림", "장비 유형을 선택하거나 새로 입력하세요.", parent=select_dialog)
                    return
    
                selected_equipment_type_id = None
                for et_id, et_name, _ in equipment_types:
                    if et_name == chosen_type_name:
                        selected_equipment_type_id = et_id
                        break
                
                if selected_equipment_type_id is None:
                    messagebox.showerror("오류", f"장비 유형 ID를 찾을 수 없습니다: {chosen_type_name}", parent=select_dialog)
                    return
    
                # Process selected items from the comparison tree
                items_to_potentially_add = [] # List of (param_name, values_tuple)
                param_names_in_selection = [] # For checking duplicates within selection itself
    
                for item_tree_id in selected_items_tree_ids:
                    values = self.comparison_tree.item(item_tree_id, "values")
                    module, part, item_name_val = values[1], values[2], values[3]
                    param_name = f"{part}_{item_name_val}" if part else item_name_val
                    if param_name not in param_names_in_selection: # Avoid processing duplicates from selection
                        items_to_potentially_add.append((param_name, values))
                        param_names_in_selection.append(param_name)
    
                # Check which of these already exist in the DB for the selected equipment type
                db_params_existing_for_type = [] # Names of params already in DB for this type
                db_params_new_for_type = []      # (param_name, values) for params not in DB for this type
    
                conn = sqlite3.connect(self.db_schema.db_path)
                cursor = conn.cursor()
                for param_name, p_values in items_to_potentially_add:
                    cursor.execute(
                        "SELECT id FROM Default_DB_Values WHERE equipment_type_id = ? AND parameter_name = ?",
                        (selected_equipment_type_id, param_name)
                    )
                    if cursor.fetchone():
                        db_params_existing_for_type.append((param_name, p_values))
                    else:
                        db_params_new_for_type.append((param_name, p_values))
                conn.close()
    
                params_for_db_operation = [] # Final list of (param_name, values) to process
    
                if db_params_existing_for_type:
                    if len(db_params_existing_for_type) == len(items_to_potentially_add):
                        messagebox.showinfo("알림", f"선택한 모든 항목이 '{chosen_type_name}' 유형으로 이미 Default DB에 존재합니다.", parent=select_dialog)
                        select_dialog.destroy()
                        return
    
                    dialog_result = messagebox.askyesnocancel(
                        "중복 항목 확인",
                        f"선택한 {len(items_to_potentially_add)}개 항목 중 {len(db_params_existing_for_type)}개는 '{chosen_type_name}' 유형으로 이미 Default DB에 존재합니다.\n\n"
                        f"- 예 (덮어쓰기): 기존 값을 새 값으로 업데이트하고, 새 항목도 추가합니다.\n"
                        f"- 아니오 (건너뛰기): 이미 존재하는 항목은 무시하고 새 항목만 추가합니다.\n"
                        f"- 취소: 작업을 취소합니다.",
                        parent=select_dialog
                    )
    
                    if dialog_result is None: # Cancel
                        select_dialog.destroy()
                        return
                    elif dialog_result: # Yes (Overwrite)
                        params_for_db_operation.extend(db_params_new_for_type)
                        params_for_db_operation.extend(db_params_existing_for_type) # Add existing ones to be overwritten
                    else: # No (Skip)
                        params_for_db_operation.extend(db_params_new_for_type)
                else: # No existing params for this type among selected items
                    params_for_db_operation.extend(db_params_new_for_type)
                
                if not params_for_db_operation:
                    messagebox.showinfo("알림", "Default DB에 추가/업데이트할 새 항목이 없습니다.", parent=select_dialog)
                    select_dialog.destroy()
                    return
    
                progress_dialog = LoadingDialog(self.window)
                progress_dialog.update_progress(0, "Default DB에 항목 추가/업데이트 중...")
                
                try:
                    num_processed = 0
                    for idx, (param_name, p_values) in enumerate(params_for_db_operation):
                        # Values from p_values: [Checkbox, Module, Part, ItemName, ...각 DB 파일 값...]
                        model_numeric_values = []
                        first_non_numeric_value = None
                        all_hyphen = True
                        calculate_minmax = is_use_auto_calc

                        for i in range(len(self.file_names)):
                            val_str = p_values[4 + i]
                            if val_str != "-":
                                all_hyphen = False
                                try:
                                    model_numeric_values.append(float(val_str))
                                except ValueError:
                                    if first_non_numeric_value is None:
                                        first_non_numeric_value = val_str
                        
                        default_val, min_val, max_val = None, None, None
    
                        if model_numeric_values:
                            default_val = sum(model_numeric_values) / len(model_numeric_values)
                            min_val = min(model_numeric_values) * 0.9
                            max_val = max(model_numeric_values) * 1.1
                        elif first_non_numeric_value is not None:
                            default_val = first_non_numeric_value
                            min_val = default_val 
                            max_val = default_val
                        elif all_hyphen:
                            self.update_log(f"Skipping '{param_name}' as all model values are '-' or invalid.")
                            continue # Skip this parameter entirely
                        else: # Should not happen if logic above is correct, but as a fallback
                            self.update_log(f"Skipping '{param_name}' due to no valid values for default.")
                            continue
    
                        # DB Operation: Add or Update
                        conn_op = sqlite3.connect(self.db_schema.db_path)
                        cursor_op = conn_op.cursor()
                        cursor_op.execute(
                            "SELECT id FROM Default_DB_Values WHERE equipment_type_id = ? AND parameter_name = ?",
                            (selected_equipment_type_id, param_name)
                        )
                        existing_db_id_tuple = cursor_op.fetchone()
                        
                        if existing_db_id_tuple: # Update
                            self.db_schema.update_default_value(
                                existing_db_id_tuple[0], param_name, default_val, min_val, max_val, conn_override=conn_op
                            )
                        else: # Add
                            self.db_schema.add_default_value(
                                selected_equipment_type_id, param_name, default_val, min_val, max_val, conn_override=conn_op
                            )
                        conn_op.commit()
                        conn_op.close()
                        num_processed += 1
                        
                        progress = (idx + 1) / len(params_for_db_operation) * 100
                        progress_dialog.update_progress(progress, f"처리 중... ({idx+1}/{len(params_for_db_operation)})")
    
                    progress_dialog.update_progress(100, "완료!")
                    progress_dialog.close()
                    
                    messagebox.showinfo(
                        "완료",
                        f"'{chosen_type_name}' 장비 유형의 Default DB에 {num_processed}개 항목이 성공적으로 추가/업데이트되었습니다.",
                        parent=select_dialog
                    )
                    
                    if hasattr(self, 'update_equipment_type_list'): # Check if Default DB tab exists
                        self.update_equipment_type_list() # Refresh Default DB tab
                    self.update_comparison_view() # Refresh comparison view
    
                except Exception as e:
                    if 'progress_dialog' in locals() and progress_dialog.top.winfo_exists():
                        progress_dialog.close()
                    messagebox.showerror("오류", f"Default DB 항목 처리 중 예외 발생:\n{str(e)}", parent=select_dialog)
                
                select_dialog.destroy()
    
            def on_cancel():
                select_dialog.destroy()
    
            ttk.Button(button_frame, text="확인", command=on_confirm).pack(side=tk.RIGHT, padx=5)
            ttk.Button(button_frame, text="취소", command=on_cancel).pack(side=tk.RIGHT, padx=5)
            
            self.window.wait_window(select_dialog)
        
            """선택한 항목을 Default DB에 추가합니다."""
            selected_items_tree_ids = self.comparison_tree.selection()
            if not selected_items_tree_ids:
                messagebox.showinfo("알림", "Default DB에 추가할 항목을 선택하세요.")
                return
    
            if not self.maint_mode:
                result = messagebox.askyesno("유지보수 모드 필요", "Default DB를 수정하려면 유지보수 모드가 필요합니다. 지금 활성화하시겠습니까?")
                if result:
                    self.toggle_maint_mode()
                else:
                    return
    
            equipment_types = self.db_schema.get_equipment_types()
            type_names = [name for _, name, _ in equipment_types]
    
            select_dialog = tk.Toplevel(self.window)
            select_dialog.title("장비 유형 선택")
            select_dialog.geometry("350x450") # Adjusted size slightly
            select_dialog.transient(self.window)
            select_dialog.grab_set()
    
            # --- UI Elements for select_dialog ---
            selected_type_var = tk.StringVar()
            list_frame = ttk.Frame(select_dialog)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            ttk.Label(list_frame, text="기존 장비 유형 선택:").pack(anchor=tk.W)
            type_listbox = tk.Listbox(list_frame, height=10, exportselection=False)
            for name_only in type_names:
                type_listbox.insert(tk.END, name_only)
            if type_names:
                type_listbox.selection_set(0)
            type_listbox.pack(fill=tk.BOTH, expand=True, pady=5)
    
            new_type_frame = ttk.Frame(select_dialog)
            new_type_frame.pack(fill=tk.X, padx=10, pady=5)
            ttk.Label(new_type_frame, text="새 장비 유형 추가:").pack(side=tk.LEFT)
            new_type_entry = ttk.Entry(new_type_frame)
            new_type_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
            
            button_frame = ttk.Frame(select_dialog)
            button_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=10, pady=10)
            # --- End UI Elements ---
    
            def on_confirm():
                nonlocal equipment_types # To allow modification if a new type is added
                chosen_type_name = ""
                
                # Determine selected or new equipment type
                if type_listbox.curselection():
                    chosen_type_name = type_listbox.get(type_listbox.curselection())
                elif new_type_entry.get().strip():
                    chosen_type_name = new_type_entry.get().strip()
                    try:
                        self.db_schema.add_equipment_type(chosen_type_name, "")
                        equipment_types = self.db_schema.get_equipment_types() # Refresh
                    except Exception as e:
                        messagebox.showerror("오류", f"새 장비 유형 추가 중 오류 발생: {str(e)}", parent=select_dialog)
                        return
                else:
                    messagebox.showinfo("알림", "장비 유형을 선택하거나 새로 입력하세요.", parent=select_dialog)
                    return
    
                selected_equipment_type_id = None
                for et_id, et_name, _ in equipment_types:
                    if et_name == chosen_type_name:
                        selected_equipment_type_id = et_id
                        break
                
                if selected_equipment_type_id is None:
                    messagebox.showerror("오류", f"장비 유형 ID를 찾을 수 없습니다: {chosen_type_name}", parent=select_dialog)
                    return
    
                # Process selected items from the comparison tree
                items_to_potentially_add = [] # List of (param_name, values_tuple)
                param_names_in_selection = [] # For checking duplicates within selection itself
    
                for item_tree_id in selected_items_tree_ids:
                    values = self.comparison_tree.item(item_tree_id, "values")
                    module, part, item_name_val = values[1], values[2], values[3]
                    param_name = f"{part}_{item_name_val}" if part else item_name_val
                    if param_name not in param_names_in_selection: # Avoid processing duplicates from selection
                        items_to_potentially_add.append((param_name, values))
                        param_names_in_selection.append(param_name)
    
                # Check which of these already exist in the DB for the selected equipment type
                db_params_existing_for_type = [] # Names of params already in DB for this type
                db_params_new_for_type = []      # (param_name, values) for params not in DB for this type
    
                conn = sqlite3.connect(self.db_schema.db_path)
                cursor = conn.cursor()
                for param_name, p_values in items_to_potentially_add:
                    cursor.execute(
                        "SELECT id FROM Default_DB_Values WHERE equipment_type_id = ? AND parameter_name = ?",
                        (selected_equipment_type_id, param_name)
                    )
                    if cursor.fetchone():
                        db_params_existing_for_type.append((param_name, p_values))
                    else:
                        db_params_new_for_type.append((param_name, p_values))
                conn.close()
    
                params_for_db_operation = [] # Final list of (param_name, values) to process
    
                if db_params_existing_for_type:
                    if len(db_params_existing_for_type) == len(items_to_potentially_add):
                        messagebox.showinfo("알림", f"선택한 모든 항목이 '{chosen_type_name}' 유형으로 이미 Default DB에 존재합니다.", parent=select_dialog)
                        select_dialog.destroy()
                        return
    
                    dialog_result = messagebox.askyesnocancel(
                        "중복 항목 확인",
                        f"선택한 {len(items_to_potentially_add)}개 항목 중 {len(db_params_existing_for_type)}개는 '{chosen_type_name}' 유형으로 이미 Default DB에 존재합니다.\n\n"
                        f"- 예 (덮어쓰기): 기존 값을 새 값으로 업데이트하고, 새 항목도 추가합니다.\n"
                        f"- 아니오 (건너뛰기): 이미 존재하는 항목은 무시하고 새 항목만 추가합니다.\n"
                        f"- 취소: 작업을 취소합니다.",
                        parent=select_dialog
                    )
    
                    if dialog_result is None: # Cancel
                        select_dialog.destroy()
                        return
                    elif dialog_result: # Yes (Overwrite)
                        params_for_db_operation.extend(db_params_new_for_type)
                        params_for_db_operation.extend(db_params_existing_for_type) # Add existing ones to be overwritten
                    else: # No (Skip)
                        params_for_db_operation.extend(db_params_new_for_type)
                else: # No existing params for this type among selected items
                    params_for_db_operation.extend(db_params_new_for_type)
                
                if not params_for_db_operation:
                    messagebox.showinfo("알림", "Default DB에 추가/업데이트할 새 항목이 없습니다.", parent=select_dialog)
                    select_dialog.destroy()
                    return
    
                progress_dialog = LoadingDialog(self.window)
                progress_dialog.update_progress(0, "Default DB에 항목 추가/업데이트 중...")
                
                try:
                    num_processed = 0
                    for idx, (param_name, p_values) in enumerate(params_for_db_operation):
                        # Values from p_values: [Checkbox, Module, Part, ItemName, ...각 DB 파일 값...]
                        model_numeric_values = []
                        first_non_numeric_value = None
                        all_hyphen = True
                        calculate_minmax = is_use_auto_calc

                        for i in range(len(self.file_names)):
                            val_str = p_values[4 + i]
                            if val_str != "-":
                                all_hyphen = False
                                try:
                                    model_numeric_values.append(float(val_str))
                                except ValueError:
                                    if first_non_numeric_value is None:
                                        first_non_numeric_value = val_str
                        
                        default_val, min_val, max_val = None, None, None
    
                        if model_numeric_values:
                            default_val = sum(model_numeric_values) / len(model_numeric_values)
                            min_val = min(model_numeric_values) * 0.9
                            max_val = max(model_numeric_values) * 1.1
                        elif first_non_numeric_value is not None:
                            default_val = first_non_numeric_value
                            min_val = default_val 
                            max_val = default_val
                        elif all_hyphen:
                            self.update_log(f"Skipping '{param_name}' as all model values are '-' or invalid.")
                            continue # Skip this parameter entirely
                        else: # Should not happen if logic above is correct, but as a fallback
                            self.update_log(f"Skipping '{param_name}' due to no valid values for default.")
                            continue
    
                        # DB Operation: Add or Update
                        conn_op = sqlite3.connect(self.db_schema.db_path)
                        cursor_op = conn_op.cursor()
                        cursor_op.execute(
                            "SELECT id FROM Default_DB_Values WHERE equipment_type_id = ? AND parameter_name = ?",
                            (selected_equipment_type_id, param_name)
                        )
                        existing_db_id_tuple = cursor_op.fetchone()
                        
                        if existing_db_id_tuple: # Update
                            self.db_schema.update_default_value(
                                existing_db_id_tuple[0], param_name, default_val, min_val, max_val, conn_override=conn_op
                            )
                        else: # Add
                            self.db_schema.add_default_value(
                                selected_equipment_type_id, param_name, default_val, min_val, max_val, conn_override=conn_op
                            )
                        conn_op.commit()
                        conn_op.close()
                        num_processed += 1
                        
                        progress = (idx + 1) / len(params_for_db_operation) * 100
                        progress_dialog.update_progress(progress, f"처리 중... ({idx+1}/{len(params_for_db_operation)})")
    
                    progress_dialog.update_progress(100, "완료!")
                    progress_dialog.close()
                    
                    messagebox.showinfo(
                        "완료",
                        f"'{chosen_type_name}' 장비 유형의 Default DB에 {num_processed}개 항목이 성공적으로 추가/업데이트되었습니다.",
                        parent=select_dialog
                    )
                    
                    if hasattr(self, 'update_equipment_type_list'): # Check if Default DB tab exists
                        self.update_equipment_type_list() # Refresh Default DB tab
                    self.update_comparison_view() # Refresh comparison view
    
                except Exception as e:
                    if 'progress_dialog' in locals() and progress_dialog.top.winfo_exists():
                        progress_dialog.close()
                    messagebox.showerror("오류", f"Default DB 항목 처리 중 예외 발생:\n{str(e)}", parent=select_dialog)
                
                select_dialog.destroy()
    
            def on_cancel():
                select_dialog.destroy()
    
            ttk.Button(button_frame, text="확인", command=on_confirm).pack(side=tk.RIGHT, padx=5)
            ttk.Button(button_frame, text="취소", command=on_cancel).pack(side=tk.RIGHT, padx=5)
            
            self.window.wait_window(select_dialog)
    
            if not self.maint_mode:
                result = messagebox.askyesno("유지보수 모드 필요", "Default DB를 수정하려면 유지보수 모드가 필요합니다. 지금 활성화하시겠습니까?")
                if result:
                    self.toggle_maint_mode()
                else:
                    return
    
            equipment_types = self.db_schema.get_equipment_types()
            type_names = [name for _, name, _ in equipment_types]
    
            select_dialog = tk.Toplevel(self.window)
            select_dialog.title("장비 유형 선택")
            select_dialog.geometry("350x450") # Adjusted size slightly
            select_dialog.transient(self.window)
            select_dialog.grab_set()
    
            # --- UI Elements for select_dialog ---
            selected_type_var = tk.StringVar()
            list_frame = ttk.Frame(select_dialog)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            ttk.Label(list_frame, text="기존 장비 유형 선택:").pack(anchor=tk.W)
            type_listbox = tk.Listbox(list_frame, height=10, exportselection=False)
            for name_only in type_names:
                type_listbox.insert(tk.END, name_only)
            if type_names:
                type_listbox.selection_set(0)
            type_listbox.pack(fill=tk.BOTH, expand=True, pady=5)
    
            new_type_frame = ttk.Frame(select_dialog)
            new_type_frame.pack(fill=tk.X, padx=10, pady=5)
            ttk.Label(new_type_frame, text="새 장비 유형 추가:").pack(side=tk.LEFT)
            new_type_entry = ttk.Entry(new_type_frame)
            new_type_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
            
            button_frame = ttk.Frame(select_dialog)
            button_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=10, pady=10)
            # --- End UI Elements ---
    
            def on_confirm():
                nonlocal equipment_types # To allow modification if a new type is added
                chosen_type_name = ""
                
                # Determine selected or new equipment type
                if type_listbox.curselection():
                    chosen_type_name = type_listbox.get(type_listbox.curselection())
                elif new_type_entry.get().strip():
                    chosen_type_name = new_type_entry.get().strip()
                    try:
                        self.db_schema.add_equipment_type(chosen_type_name, "")
                        equipment_types = self.db_schema.get_equipment_types() # Refresh
                    except Exception as e:
                        messagebox.showerror("오류", f"새 장비 유형 추가 중 오류 발생: {str(e)}", parent=select_dialog)
                        return
                else:
                    messagebox.showinfo("알림", "장비 유형을 선택하거나 새로 입력하세요.", parent=select_dialog)
                    return
    
                selected_equipment_type_id = None
                for et_id, et_name, _ in equipment_types:
                    if et_name == chosen_type_name:
                        selected_equipment_type_id = et_id
                        break
                
                if selected_equipment_type_id is None:
                    messagebox.showerror("오류", f"장비 유형 ID를 찾을 수 없습니다: {chosen_type_name}", parent=select_dialog)
                    return
    
                # Process selected items from the comparison tree
                items_to_potentially_add = [] # List of (param_name, values_tuple)
                param_names_in_selection = [] # For checking duplicates within selection itself
    
                for item_tree_id in selected_items_tree_ids:
                    values = self.comparison_tree.item(item_tree_id, "values")
                    module, part, item_name_val = values[1], values[2], values[3]
                    param_name = f"{part}_{item_name_val}" if part else item_name_val
                    if param_name not in param_names_in_selection: # Avoid processing duplicates from selection
                        items_to_potentially_add.append((param_name, values))
                        param_names_in_selection.append(param_name)
    
                # Check which of these already exist in the DB for the selected equipment type
                db_params_existing_for_type = [] # Names of params already in DB for this type
                db_params_new_for_type = []      # (param_name, values) for params not in DB for this type
    
                conn = sqlite3.connect(self.db_schema.db_path)
                cursor = conn.cursor()
                for param_name, p_values in items_to_potentially_add:
                    cursor.execute(
                        "SELECT id FROM Default_DB_Values WHERE equipment_type_id = ? AND parameter_name = ?",
                        (selected_equipment_type_id, param_name)
                    )
                    if cursor.fetchone():
                        db_params_existing_for_type.append((param_name, p_values))
                    else:
                        db_params_new_for_type.append((param_name, p_values))
                conn.close()
    
                params_for_db_operation = [] # Final list of (param_name, values) to process
    
                if db_params_existing_for_type:
                    if len(db_params_existing_for_type) == len(items_to_potentially_add):
                        messagebox.showinfo("알림", f"선택한 모든 항목이 '{chosen_type_name}' 유형으로 이미 Default DB에 존재합니다.", parent=select_dialog)
                        select_dialog.destroy()
                        return
    
                    dialog_result = messagebox.askyesnocancel(
                        "중복 항목 확인",
                        f"선택한 {len(items_to_potentially_add)}개 항목 중 {len(db_params_existing_for_type)}개는 '{chosen_type_name}' 유형으로 이미 Default DB에 존재합니다.\n\n"
                        f"- 예 (덮어쓰기): 기존 값을 새 값으로 업데이트하고, 새 항목도 추가합니다.\n"
                        f"- 아니오 (건너뛰기): 이미 존재하는 항목은 무시하고 새 항목만 추가합니다.\n"
                        f"- 취소: 작업을 취소합니다.",
                        parent=select_dialog
                    )
    
                    if dialog_result is None: # Cancel
                        select_dialog.destroy()
                        return
                    elif dialog_result: # Yes (Overwrite)
                        params_for_db_operation.extend(db_params_new_for_type)
                        params_for_db_operation.extend(db_params_existing_for_type) # Add existing ones to be overwritten
                    else: # No (Skip)
                        params_for_db_operation.extend(db_params_new_for_type)
                else: # No existing params for this type among selected items
                    params_for_db_operation.extend(db_params_new_for_type)
                
                if not params_for_db_operation:
                    messagebox.showinfo("알림", "Default DB에 추가/업데이트할 새 항목이 없습니다.", parent=select_dialog)
                    select_dialog.destroy()
                    return
    
                progress_dialog = LoadingDialog(self.window)
                progress_dialog.update_progress(0, "Default DB에 항목 추가/업데이트 중...")
                
                try:
                    num_processed = 0
                    for idx, (param_name, p_values) in enumerate(params_for_db_operation):
                        # Values from p_values: [Checkbox, Module, Part, ItemName, ...각 DB 파일 값...]
                        model_numeric_values = []
                        first_non_numeric_value = None
                        all_hyphen = True
                        calculate_minmax = is_use_auto_calc

                        for i in range(len(self.file_names)):
                            val_str = p_values[4 + i]
                            if val_str != "-":
                                all_hyphen = False
                                try:
                                    model_numeric_values.append(float(val_str))
                                except ValueError:
                                    if first_non_numeric_value is None:
                                        first_non_numeric_value = val_str
                        
                        default_val, min_val, max_val = None, None, None
    
                        if model_numeric_values:
                            default_val = sum(model_numeric_values) / len(model_numeric_values)
                            min_val = min(model_numeric_values) * 0.9
                            max_val = max(model_numeric_values) * 1.1
                        elif first_non_numeric_value is not None:
                            default_val = first_non_numeric_value
                            min_val = default_val 
                            max_val = default_val
                        elif all_hyphen:
                            self.update_log(f"Skipping '{param_name}' as all model values are '-' or invalid.")
                            continue # Skip this parameter entirely
                        else: # Should not happen if logic above is correct, but as a fallback
                            self.update_log(f"Skipping '{param_name}' due to no valid values for default.")
                            continue
    
                        # DB Operation: Add or Update
                        conn_op = sqlite3.connect(self.db_schema.db_path)
                        cursor_op = conn_op.cursor()
                        cursor_op.execute(
                            "SELECT id FROM Default_DB_Values WHERE equipment_type_id = ? AND parameter_name = ?",
                            (selected_equipment_type_id, param_name)
                        )
                        existing_db_id_tuple = cursor_op.fetchone()
                        
                        if existing_db_id_tuple: # Update
                            self.db_schema.update_default_value(
                                existing_db_id_tuple[0], param_name, default_val, min_val, max_val, conn_override=conn_op
                            )
                        else: # Add
                            self.db_schema.add_default_value(
                                selected_equipment_type_id, param_name, default_val, min_val, max_val, conn_override=conn_op
                            )
                        conn_op.commit()
                        conn_op.close()
                        num_processed += 1
                        
                        progress = (idx + 1) / len(params_for_db_operation) * 100
                        progress_dialog.update_progress(progress, f"처리 중... ({idx+1}/{len(params_for_db_operation)})")
    
                    progress_dialog.update_progress(100, "완료!")
                    progress_dialog.close()
                    
                    messagebox.showinfo(
                        "완료",
                        f"'{chosen_type_name}' 장비 유형의 Default DB에 {num_processed}개 항목이 성공적으로 추가/업데이트되었습니다.",
                        parent=select_dialog
                    )
                    
                    if hasattr(self, 'update_equipment_type_list'): # Check if Default DB tab exists
                        self.update_equipment_type_list() # Refresh Default DB tab
                    self.update_comparison_view() # Refresh comparison view
    
                except Exception as e:
                    if 'progress_dialog' in locals() and progress_dialog.top.winfo_exists():
                        progress_dialog.close()
                    messagebox.showerror("오류", f"Default DB 항목 처리 중 예외 발생:\n{str(e)}", parent=select_dialog)
                
                select_dialog.destroy()
    
            def on_cancel():
                select_dialog.destroy()
    
            ttk.Button(button_frame, text="확인", command=on_confirm).pack(side=tk.RIGHT, padx=5)
            ttk.Button(button_frame, text="취소", command=on_cancel).pack(side=tk.RIGHT, padx=5)
            
            self.window.wait_window(select_dialog)
    def add_to_default_db(self):
        """체크/선택된 항목을 Default DB에 일괄 저장(트랜잭션) 및 변경 이력 기록, 상세 피드백 제공"""
        import pandas as pd
        if not self.maint_mode:
            result = messagebox.askyesno(
                "유지보수 모드 필요",
                "Default DB를 수정하려면 유지보수 모드가 필요합니다. 지금 활성화하시겠습니까?")
            if result:
                self.toggle_maint_mode()
            else:
                return

        checked_items = [item_id for item_id in self.comparison_tree.get_children()
                         if self.comparison_tree.item(item_id, "values") and self.comparison_tree.item(item_id, "values")[0] == "☑"]
        selected_items_tree_ids = checked_items if checked_items else self.comparison_tree.selection()
        if not selected_items_tree_ids:
            messagebox.showinfo("알림", "Default DB에 추가할 항목을 체크하거나 선택하세요.")
            return
        equipment_types = self.db_schema.get_equipment_types()
        type_names = [name for _, name, _ in equipment_types]

        select_dialog = tk.Toplevel(self.window)
        select_dialog.title("장비 유형 선택")
        select_dialog.geometry("400x600")
        select_dialog.transient(self.window)
        select_dialog.grab_set()
        selected_type_var = tk.StringVar()
        list_frame = ttk.Frame(select_dialog)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        ttk.Label(list_frame, text="기존 장비 유형 선택:").pack(anchor=tk.W)
        type_listbox = tk.Listbox(list_frame, height=10, exportselection=False)
        for name_only in type_names:
            type_listbox.insert(tk.END, name_only)
        if type_names:
            type_listbox.selection_set(0)
        type_listbox.pack(fill=tk.BOTH, expand=True, pady=5)
        options_frame = ttk.LabelFrame(select_dialog, text="옵션")
        options_frame.pack(fill=tk.X, padx=10, pady=5)
        auto_overwrite_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="중복 항목 자동 덧어쓰기", variable=auto_overwrite_var).pack(anchor=tk.W, padx=5, pady=3)
        use_auto_calc_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="자동 계산 값 사용 (min/max)", variable=use_auto_calc_var).pack(anchor=tk.W, padx=5, pady=3)
        new_type_frame = ttk.Frame(select_dialog)
        new_type_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(new_type_frame, text="새 장비 유형 추가:").pack(side=tk.LEFT)
        new_type_entry = ttk.Entry(new_type_frame)
        new_type_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        button_frame = ttk.Frame(select_dialog)
        button_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=10, pady=10)

        def extract_param_name(values):
            part, item_name_val = values[2], values[3]
            return f"{part}_{item_name_val}" if part else item_name_val

        def on_confirm():
            nonlocal equipment_types
            chosen_type_name = ""
            is_auto_overwrite = auto_overwrite_var.get()
            is_use_auto_calc = use_auto_calc_var.get()
            if type_listbox.curselection():
                chosen_type_name = type_listbox.get(type_listbox.curselection())
            elif new_type_entry.get().strip():
                chosen_type_name = new_type_entry.get().strip()
                try:
                    self.db_schema.add_equipment_type(chosen_type_name, "")
                    equipment_types = self.db_schema.get_equipment_types()
                except Exception as e:
                    messagebox.showerror("오류", f"새 장비 유형 추가 중 오류 발생: {str(e)}", parent=select_dialog)
                    return
            else:
                messagebox.showinfo("알림", "장비 유형을 선택하거나 새로 입력하세요.", parent=select_dialog)
                return
            selected_equipment_type_id = None
            for et_id, et_name, _ in equipment_types:
                if et_name == chosen_type_name:
                    selected_equipment_type_id = et_id
                    break
            if selected_equipment_type_id is None:
                messagebox.showerror("오류", f"장비 유형 ID를 찾을 수 없습니다: {chosen_type_name}", parent=select_dialog)
                return
            # 1. 선택 항목 정리 및 중복 제거
            items_to_add = []
            seen_param_names = set()
            for item_tree_id in selected_items_tree_ids:
                values = self.comparison_tree.item(item_tree_id, "values")
                param_name = extract_param_name(values)
                if param_name not in seen_param_names:
                    items_to_add.append((param_name, values))
                    seen_param_names.add(param_name)
            # 2. DB 중복 체크
            db_params_existing, db_params_new = [], []
            with self.db_schema.get_connection() as conn:
                cursor = conn.cursor()
                for param_name, p_values in items_to_add:
                    cursor.execute(
                        "SELECT id, default_value, min_spec, max_spec FROM Default_DB_Values WHERE equipment_type_id = ? AND parameter_name = ?",
                        (selected_equipment_type_id, param_name)
                    )
                    row = cursor.fetchone()
                    if row:
                        db_params_existing.append((param_name, p_values, row))
                    else:
                        db_params_new.append((param_name, p_values, None))
            # 3. 덮어쓰기/스킵 옵션 처리
            params_for_db_op = []
            if db_params_existing:
                if len(db_params_existing) == len(items_to_add):
                    messagebox.showinfo("알림", f"선택한 모든 항목이 '{chosen_type_name}' 유형으로 이미 Default DB에 존재합니다.", parent=select_dialog)
                    select_dialog.destroy()
                    return
                if is_auto_overwrite:
                    params_for_db_op.extend(db_params_new)
                    params_for_db_op.extend(db_params_existing)
                else:
                    dialog_result = messagebox.askyesnocancel(
                        "중복 항목 확인",
                        f"선택한 {len(items_to_add)}개 항목 중 {len(db_params_existing)}개는 이미 존재합니다.\n\n"
                        "- 예 (덮어쓰기): 기존 값을 새 값으로 업데이트\n"
                        "- 아니오 (건너뛰기): 이미 존재하는 항목은 무시하고 새 항목만 추가\n"
                        "- 취소: 작업을 취소합니다.",
                        parent=select_dialog
                    )
                    if dialog_result is None:
                        select_dialog.destroy()
                        return
                    elif dialog_result:
                        params_for_db_op.extend(db_params_new)
                        params_for_db_op.extend(db_params_existing)
                    else:
                        params_for_db_op.extend(db_params_new)
            else:
                params_for_db_op.extend(db_params_new)
            
            # ... (나머지 코드는 동일합니다)

    def create_report_tab(self):
        # 보고서 탭 생성
        report_frame = ttk.Frame(self.comparison_notebook)
        self.comparison_notebook.add(report_frame, text="📊 비교 보고서")
        
        # 통계 프레임과 차트 프레임을 나란히 배치
        stats_frame = ttk.LabelFrame(report_frame, text="📈 비교 통계")
        stats_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        chart_frame = ttk.LabelFrame(report_frame, text="📊 시각화")
        chart_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        if self.merged_df is not None:
            # 통계 정보 계산
            stats = self.calculate_stats()
            
            # 전체 통계
            total_frame = ttk.LabelFrame(stats_frame, text="전체 통계")
            total_frame.pack(fill=tk.X, padx=5, pady=5)
            ttk.Label(total_frame, 
                     text=f"총 {stats['total_diff']}개 항목에서 차이가 발견되었습니다.",
                     font=('Arial', 10, 'bold')).pack(pady=5)
            
            # 모듈별 통계
            module_frame = ttk.LabelFrame(stats_frame, text="모듈별 차이")
            module_frame.pack(fill=tk.X, padx=5, pady=5)
            
            # 모듈별 트리뷰
            module_tree = ttk.Treeview(module_frame, height=10)
            module_tree["columns"] = ["count"]
            module_tree.heading("#0", text="모듈", anchor="w")
            module_tree.heading("count", text="차이 개수", anchor="w")
            module_tree.column("#0", width=200)
            module_tree.column("count", width=100)
            
            for module, count in sorted(stats['module_diff'].items(), key=lambda x: x[1], reverse=True):
                module_tree.insert("", "end", text=module, values=(count,))
            
            module_tree.pack(fill=tk.X, padx=5, pady=5)
            
            # 파트별 통계
            part_frame = ttk.LabelFrame(stats_frame, text="상세 경로별 차이")
            part_frame.pack(fill=tk.X, padx=5, pady=5)
            
            # 파트별 트리뷰
            part_tree = ttk.Treeview(part_frame, height=15)
            part_tree["columns"] = ["count"]
            part_tree.heading("#0", text="경로", anchor="w")
            part_tree.heading("count", text="차이 개수", anchor="w")
            part_tree.column("#0", width=200)
            part_tree.column("count", width=100)
            
            for part, count in sorted(stats['part_diff'].items(), key=lambda x: x[1], reverse=True):
                part_tree.insert("", "end", text=part, values=(count,))
            
            part_tree.pack(fill=tk.X, padx=5, pady=5)
            
            # 차트 생성
            try:
                # 한글 폰트 설정
                plt.rcParams['font.family'] = 'Malgun Gothic'
                
                # 그래프 생성
                fig = plt.figure(figsize=(8, 10))
                
                # 모듈별 차이 그래프
                ax1 = fig.add_subplot(211)
                modules = list(stats['module_diff'].keys())
                counts = list(stats['module_diff'].values())
                bars1 = ax1.bar(range(len(modules)), counts)
                ax1.set_title('Module Differences')
                ax1.set_xticks(range(len(modules)))
                ax1.set_xticklabels(modules, rotation=45, ha='right')
                
                # 막대 위에 값 표시
                for bar in bars1:
                    height = bar.get_height()
                    ax1.text(bar.get_x() + bar.get_width()/2., height,
                            f'{int(height)}',
                            ha='center', va='bottom')
                
                # 파트별 차이 그래프 (상위 10개만)
                ax2 = fig.add_subplot(212)
                parts = list(stats['part_diff'].keys())[:10]
                part_counts = list(stats['part_diff'].values())[:10]
                bars2 = ax2.bar(range(len(parts)), part_counts)
                ax2.set_title('Top 10 Path Differences')
                ax2.set_xticks(range(len(parts)))
                ax2.set_xticklabels(parts, rotation=45, ha='right')
                
                # 막대 위에 값 표시
                for bar in bars2:
                    height = bar.get_height()
                    ax2.text(bar.get_x() + bar.get_width()/2., height,
                            f'{int(height)}',
                            ha='center', va='bottom')
                
                plt.tight_layout()
                
                # 캔버스에 그래프 표시
                canvas = FigureCanvasTkAgg(fig, master=chart_frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
                
            except Exception as e:
                error_msg = f"차트 생성 중 오류가 발생했습니다: {str(e)}"
                ttk.Label(chart_frame, text=error_msg, wraplength=300).pack(pady=20)
                print(error_msg)

    def calculate_stats(self):
        stats = {}
        grouped = self.merged_df.groupby(["Module", "Part", "ItemName"])
        
        # 모듈별 차이 개수
        module_diff = {}
        # 파트별 차이 개수
        part_diff = {}
        # 전체 차이 개수
        total_diff = 0
        
        for (module, part, item_name), group in grouped:
            values = [group[group["Model"] == model]["ItemValue"].values[0] 
                     if len(group[group["Model"] == model]["ItemValue"]) > 0 
                     else "-" for model in self.file_names]
            
            if len(set(values)) > 1:  # 값이 다른 경우
                total_diff += 1
                module_diff[module] = module_diff.get(module, 0) + 1
                part_key = f"{module} > {part}"
                part_diff[part_key] = part_diff.get(part_key, 0) + 1
        
        return {
            'total_diff': total_diff,
            'module_diff': module_diff,
            'part_diff': part_diff
        }

    def on_qc_equipment_type_selected(self, event=None):
        print("QC Tab: Equipment type selected")
        # TODO: 실제 로직 구현 필요
        pass

    def create_about_tab(self):
        # About 탭 생성
        self.about_frame = ttk.Frame(self.main_notebook)
        self.main_notebook.add(self.about_frame, text="About")
        
        # 스타일 설정
        style = ttk.Style()
        style.configure("Title.TLabel", font=('Helvetica', 16, 'bold'))
        style.configure("Header.TLabel", font=('Helvetica', 12, 'bold'))
        style.configure("Content.TLabel", font=('Helvetica', 10))
        
        # 컨테이너 프레임
        container = ttk.Frame(self.about_frame, padding="20")
        container.pack(expand=True, fill=tk.BOTH)
        
        # 프로그램 제목
        title_frame = ttk.Frame(container)
        title_frame.pack(fill=tk.X, pady=(0, 20))
        ttk.Label(title_frame, text="DB 관리 프로그램", style="Title.TLabel").pack()
        
        # 정보 섹션들
        sections = [
            ("Product Information", [
                ("Version", "1.0.0"),
                ("Release Date", "2025-02-04"),
            ]),
            ("Development", [
                ("Developer", "Levi Beak / 백광림"),
                ("Organization", "Quality Assurance Team"),
                ("Contact", "levi.beak@parksystems.com"),
            ]),
                        ]
        
        for section_title, items in sections:
            # 섹션 프레임
            section_frame = ttk.LabelFrame(container, text=section_title, padding="10")
            section_frame.pack(fill=tk.X, pady=(0, 10))
            
            # 그리드 구성
            for i, (key, value) in enumerate(items):
                ttk.Label(section_frame, text=key + ":", style="Header.TLabel").grid(
                    row=i, column=0, sticky="w", padx=(0, 10), pady=2)
                ttk.Label(section_frame, text=value, style="Content.TLabel").grid(
                    row=i, column=1, sticky="w", pady=2)
        
        # 프로그램 설명
        desc_frame = ttk.LabelFrame(container, text="Description", padding="10")
        desc_frame.pack(fill=tk.X, pady=(0, 10))
        
        description = """이 프로그램은 XES 데이터베이스 비교 및 관리를 위한 프로그램입니다.
        
주요 기능:
• 다중 DB 파일 비교 분석
• 차이점 자동 감지 및 하이라이트
• 상세 비교 보고서 생성
• 데이터 시각화 및 통계 분석
• QC 스펙 관리 및 검증(추가 예정)
"""
        
        ttk.Label(desc_frame, text=description, style="Content.TLabel", 
                 wraplength=500, justify="left").pack(anchor="w")

    def show_about(self):
        # About 창 생성
        about_window = tk.Toplevel(self.window)
        about_window.title("About")
        about_window.geometry("600x800")
        
        # 스타일 설정
        style = ttk.Style()
        style.configure("Title.TLabel", font=('Helvetica', 16, 'bold'))
        style.configure("Header.TLabel", font=('Helvetica', 12, 'bold'))
        style.configure("Content.TLabel", font=('Helvetica', 10))
        
        # 컨테이너 프레임
        container = ttk.Frame(about_window, padding="20")
        container.pack(expand=True, fill=tk.BOTH)
        
        # 프로그램 제목
        title_frame = ttk.Frame(container)
        title_frame.pack(fill=tk.X, pady=(0, 20))
        ttk.Label(title_frame, text="DB 관리 프로그램", style="Title.TLabel").pack()
        
        # 정보 섹션들
        sections = [
            ("Product Information", [
                ("Version", "1.0.0"),
                ("Release Date", "2025-02-04"),
            ]),
            ("Development", [
                ("Developer", "Levi Beak / 백광림"),
                ("Organization", "Quality Assurance Team"),
                ("Contact", "levi.beak@parksystems.com"),
            ]),
        ]
        
        for section_title, items in sections:
            # 섹션 프레임
            section_frame = ttk.LabelFrame(container, text=section_title, padding="10")
            section_frame.pack(fill=tk.X, pady=(0, 10))
            
            # 그리드 구성
            for i, (key, value) in enumerate(items):
                ttk.Label(section_frame, text=key + ":", style="Header.TLabel").grid(
                    row=i, column=0, sticky="w", padx=(0, 10), pady=2)
                ttk.Label(section_frame, text=value, style="Content.TLabel").grid(
                    row=i, column=1, sticky="w", pady=2)
        
        # 프로그램 설명
        desc_frame = ttk.LabelFrame(container, text="Description", padding="10")
        desc_frame.pack(fill=tk.X, pady=(0, 10))
        
        description = """이 프로그램은 XES 데이터베이스 비교 및 관리를 위한 프로그램입니다.
        
주요 기능:
• 다중 DB 파일 비교 분석
• 차이점 자동 감지 및 하이라이트
• 상세 비교 보고서 생성
• 데이터 시각화 및 통계 분석
• QC 스펙 관리 및 검증(추가 예정)
"""
        
        ttk.Label(desc_frame, text=description, style="Content.TLabel", 
                 wraplength=500, justify="left").pack(anchor="w")
        
        # 리비전 히스토리 데이터
        revisions = [
            {
                "version": "1.0.0",
                "date": "2025-02-04",
                "summary": "초기 버전 출시",
                "details": {
                    "Features": [
                        "다중 XES DB 파일 비교 분석 기능",
                        "자동 차이점 감지 및 하이라이트",
                        "상세 비교 보고서 생성"
                    ],
                    "Improvements": [
                        "데이터 시각화 도구 추가"
                    ],
                    "Bug Fixes": [
                        "파일 로드 시 인코딩 문제 수정",
                        "메모리 사용량 최적화"
                    ]
                }
            }
            # 새로운 버전이 출시될 때마다 여기에 추가
        ]
        
        # 리비전 히스토리를 위한 트리뷰 생성
        revision_frame = ttk.LabelFrame(container, text="Revision History", padding="10")
        revision_frame.pack(fill=tk.X, pady=(0, 10))
        
        revision_tree = ttk.Treeview(revision_frame, height=6)
        revision_tree["columns"] = ("Version", "Date", "Summary")
        revision_tree.heading("#0", text="")
        revision_tree.column("#0", width=0, stretch=False)
        
        for col, width in [("Version", 70), ("Date", 100), ("Summary", 400)]:
            revision_tree.heading(col, text=col)
            revision_tree.column(col, width=width)
        
        # 리비전 데이터 추가
        for rev in revisions:
            revision_tree.insert("", 0, values=(
                rev["version"],
                rev["date"],
                rev["summary"]
            ), tags=("revision",))
        
        # 더블 클릭 이벤트 처리
        def show_revision_details(event):
            item = revision_tree.selection()[0]
            version = revision_tree.item(item)["values"][0]
            
            # 해당 버전의 상세 정보 찾기
            rev_info = next(r for r in revisions if r["version"] == version)
            
            # 상세 정보 창 생성
            detail_window = tk.Toplevel(about_window)
            detail_window.title(f"Version {version} Details")
            detail_window.geometry("500x800")  # About 창과 같은 높이로 설정
            detail_window.transient(about_window)
            detail_window.grab_set()
            
            # About 창 오른쪽에 나란히 표시 (화면 범위 체크 추가)
            about_x = about_window.winfo_x()
            about_y = about_window.winfo_y()
            about_width = about_window.winfo_width()
            
            # 화면 크기 확인
            screen_width = detail_window.winfo_screenwidth()
            
            # 새 창의 x 좌표 계산
            new_x = about_x + about_width + 10
            
            # 화면 밖으로 넘어갈 경우 About 창 왼쪽에 표시
            if new_x + 500 > screen_width:  # 500은 detail_window의 너비
                new_x = about_x - 510  # 왼쪽에 표시 (간격 10픽셀)
            
            detail_window.geometry(f"500x800+{new_x}+{about_y}")
            
            # 스타일 설정
            style = ttk.Style()
            style.configure("Category.TLabel", font=('Helvetica', 11, 'bold'))
            style.configure("Item.TLabel", font=('Helvetica', 10))
            
            # 컨테이너 생성
            detail_container = ttk.Frame(detail_window, padding="20")
            detail_container.pack(fill=tk.BOTH, expand=True)
            
            # 버전 정보 헤더
            ttk.Label(
                detail_container,
                text=f"Version {version} ({rev_info['date']})",
                style="Title.TLabel"
            ).pack(anchor="w", pady=(0, 20))
            
            # 카테고리별 상세 정보 표시
            for category, items in rev_info["details"].items():
                # 카테고리 제목
                ttk.Label(
                    detail_container,
                    text=category,
                    style="Category.TLabel"
                ).pack(anchor="w", pady=(10, 5))
                
                # 항목들
                for item in items:
                    ttk.Label(
                        detail_container,
                        text=f"• {item}",
                        style="Item.TLabel",
                        wraplength=450
                    ).pack(anchor="w", padx=(20, 0))
            
            # 닫기 버튼
            ttk.Button(
                detail_container,
                text="닫기",
                command=detail_window.destroy
            ).pack(pady=(20, 0))
        
        # 더블 클릭 이벤트 바인딩
        revision_tree.bind("<Double-1>", show_revision_details)
        
        # 스크롤바 추가
        scrollbar = ttk.Scrollbar(revision_frame, orient="vertical", command=revision_tree.yview)
        revision_tree.configure(yscrollcommand=scrollbar.set)
        
        # 레이아웃
        revision_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 닫기 버튼
        ttk.Button(container, text="닫기", command=about_window.destroy).pack(pady=(0, 10))

    def show_user_guide(self, event=None):
        print("사용 설명서가 호출되었습니다. (F1 키 또는 메뉴 선택)")
        guide_window = tk.Toplevel(self.window)
        guide_window.title("DB 관리 도구 사용 설명서")
        guide_window.geometry("800x600")
        guide_window.resizable(True, True)  # 창 크기 조절 가능
        
        # 부모 창 중앙에 위치
        x = self.window.winfo_x() + (self.window.winfo_width() // 2) - (800 // 2)
        y = self.window.winfo_y() + (self.window.winfo_height() // 2) - (600 // 2)
        guide_window.geometry(f"800x600+{x}+{y}")
        
        # 스타일 설정
        style = ttk.Style()
        style.configure("Title.TLabel", font=('Helvetica', 16, 'bold'))
        style.configure("Heading.TLabel", font=('Helvetica', 12, 'bold'))
        style.configure("Content.TLabel", font=('Helvetica', 10))
        
        # 메인 프레임과 캔버스, 스크롤바 설정
        main_frame = ttk.Frame(guide_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 내용 구성
        sections = [
            {
                "title": "시작하기",
                "content": [
                    "1. 프로그램 실행 후 '파일' 메뉴에서 '폴더 열기' 선택",
                    "2. DB Editor에서 Export한 .txt 파일이 있는 폴더 선택",
                    "3. 최대 6개의 DB 파일들을 선택하여 비교 분석 실행"
                ]
            },
            {
                "title": "주요 기능",
                "content": [
                    "• DB 파일 비교 분석",
                    "  - 여러 DB 파일의 내용을 자동으로 비교",
                    "  - 차이점 자동 감지 및 하이라이트",
                    "  - 상세 비교 결과 제공",
                    "",
                    "• QC 검수 기능 (추가 예정)",
                    "  - 설정된 기준에 따른 자동 검증",
                    "  - 오류 항목 자동 감지",
                    "  - 검수 결과 리포트 생성"
                ]
            },
            {
                "title": "단축키",
                "content": [
                    "• Ctrl + O : 폴더 열기",
                    "• Ctrl + Q : 프로그램 종료",
                    "• F1 : 도움말 열기"
                ]
            },
            {
                "title": "자주 묻는 질문",
                "content": [
                    "Q: 지원하는 파일 형식은 무엇인가요?",
                    "A: DB Editor에서 Export한 .txt 파일을 지원합니다.",
                    "",
                    "Q: 한 번에 몇 개의 파일을 비교할 수 있나요?",
                    "A: 최대 6개의 파일을 동시에 비교할 수 있습니다.",
                    ""
                ]
            }
        ]
        
        # 제목
        ttk.Label(
            scrollable_frame,
            text="DB 관리 프로그램 사용 설명서",
            style="Title.TLabel"
        ).pack(pady=(0, 20))
        
        # 섹션별 내용 추가
        for section in sections:
            # 섹션 제목
            ttk.Label(
                scrollable_frame,
                text=section["title"],
                style="Heading.TLabel"
            ).pack(anchor="w", pady=(15, 5))
            
            # 섹션 내용
            for line in section["content"]:
                ttk.Label(
                    scrollable_frame,
                    text=line,
                    style="Content.TLabel",
                    wraplength=700,
                    justify="left"
                ).pack(anchor="w", padx=(20, 0))
        
        # 레이아웃 설정
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def update_log(self, message):
        if hasattr(self, 'log_text'):
            self.log_text.insert("end", message + "\n")
            self.log_text.see("end")

    def export_report(self):
        """현재 로드된 DB 파일에 대한 보고서를 Excel 형식으로 내보냅니다."""
        if self.merged_df is None or len(self.file_names) == 0:
            messagebox.showinfo("알림", "보고서를 생성하기 위한 DB 파일을 먼저 로드하세요.")
            return
        
        # 저장 파일 선택
        file_path = filedialog.asksaveasfilename(
            title="보고서 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
            initialfile=f"DB_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file_path:
            return
        
        try:
            # 로딩 다이얼로그 표시
            loading_dialog = LoadingDialog(self.window)
            loading_dialog.update_progress(0, "보고서 생성 중...")
            
            # Excel 워크북 생성
            wb = openpyxl.Workbook()
            
            # 기본 시트 이름 변경
            summary_sheet = wb.active
            summary_sheet.title = "요약"
            
            # 헤더 스타일 설정
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 테두리 스타일
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 요약 정보 작성
            summary_sheet['A1'] = "DB 비교 보고서"
            summary_sheet['A1'].font = Font(name="맑은 고딕", size=16, bold=True)
            summary_sheet.merge_cells('A1:D1')
            
            summary_sheet['A3'] = "생성 일시:"
            summary_sheet['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            summary_sheet['A4'] = "로드된 파일 수:"
            summary_sheet['B4'] = len(self.file_names)
            
            summary_sheet['A5'] = "로드된 파일:"
            summary_sheet['B5'] = ", ".join(self.file_names)
            
            # 진행상태 업데이트
            loading_dialog.update_progress(20, "요약 시트 생성 중...")
            
            # 모델별 DB_ItemName 요약 테이블
            summary_sheet['A7'] = "모델별 DB_ItemName 요약"
            summary_sheet['A7'].font = Font(name="맑은 고딕", size=12, bold=True)
            summary_sheet.merge_cells('A7:D7')
            
            # 헤더 행
            summary_sheet['A9'] = "모델"
            summary_sheet['B9'] = "DB_ItemName 수"
            
            for col in ['A9', 'B9']:
                summary_sheet[col].font = header_font
                summary_sheet[col].fill = header_fill
                summary_sheet[col].alignment = header_alignment
                summary_sheet[col].border = thin_border
            
            # 모델별 DB_ItemName 수 집계
            row = 10
            for model in self.file_names:
                model_params = self.merged_df[self.merged_df['Model'] == model]['Parameter'].nunique()
                
                summary_sheet[f'A{row}'] = model
                summary_sheet[f'B{row}'] = model_params
                
                for col in ['A', 'B']:
                    summary_sheet[f'{col}{row}'].border = thin_border
                
                row += 1
            
            # 열 너비 조정
            summary_sheet.column_dimensions['A'].width = 30
            summary_sheet.column_dimensions['B'].width = 15
            
            # 진행상태 업데이트
            loading_dialog.update_progress(40, "상세 데이터 시트 생성 중...")
            
            # 상세 데이터 시트 생성
            detail_sheet = wb.create_sheet(title="상세 데이터")
            
            # 열 제목 설정
            headers = ["DB_ItemName"] + self.file_names
            for col, header in enumerate(headers, 1):
                cell = detail_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 모든 DB_ItemName 목록 얻기
            all_params = sorted(self.merged_df['DB_ItemName'].unique())
            
            # 진행상태 업데이트
            loading_dialog.update_progress(60, "데이터 입력 중...")
            
            # 각 DB_ItemName별 모델 값 기입
            for row_idx, param in enumerate(all_params, 2):
                # 25개 DB_ItemName마다 진행상태 업데이트
                if row_idx % 25 == 0:
                    progress = 60 + (row_idx / len(all_params)) * 30  # 60%에서 90%까지
                    loading_dialog.update_progress(progress, f"데이터 입력 중... ({row_idx}/{len(all_params)})")
                
                # DB_ItemName 이름
                detail_sheet.cell(row=row_idx, column=1, value=param).border = thin_border
                
                # 각 모델별 값
                for col_idx, model in enumerate(self.file_names, 2):
                    value_df = self.merged_df[(self.merged_df['Model'] == model) & 
                                        (self.merged_df['DB_ItemName'] == param)]
                    
                    if not value_df.empty:
                        value = value_df.iloc[0]['Value']
                        cell = detail_sheet.cell(row=row_idx, column=col_idx, value=value)
                    else:
                        cell = detail_sheet.cell(row=row_idx, column=col_idx, value="-")
                    
                    cell.border = thin_border
            
            # 열 너비 조정
            for col in range(1, len(headers) + 1):
                column_letter = openpyxl.utils.get_column_letter(col)
                detail_sheet.column_dimensions[column_letter].width = 20
            
            # 셀 스타일 조정
            detail_sheet.freeze_panes = 'B2'  # 첫 번째 행과 열 고정
            
            # 진행상태 업데이트
            loading_dialog.update_progress(95, "파일 저장 중...")
            
            # 파일 저장
            wb.save(file_path)
            
            loading_dialog.update_progress(100, "완료!")
            loading_dialog.close()
            
            # 완료 메시지
            messagebox.showinfo("보고서 생성 완료", f"보고서가 성공적으로 생성되었습니다.\n경로: {file_path}")
            
            # 로그 업데이트
            self.update_log(f"보고서 내보내기: {file_path}")
            
            # 파일 열기 확인
            if messagebox.askyesno("파일 열기", "생성된 보고서를 바로 열어보시겠습니까?"):
                import os
                os.startfile(file_path)
        
        except Exception as e:
            if 'loading_dialog' in locals():
                loading_dialog.close()
            messagebox.showerror("오류", f"보고서 생성 중 오류가 발생했습니다:\n{str(e)}")
            self.update_log(f"보고서 생성 오류: {str(e)}")
        
class LoadingDialog:
    def __init__(self, parent):
        self.top = tk.Toplevel(parent)
        self.top.title("로딩 중...")
        
        # 부모 창 중앙에 위치
        window_width = 300
        window_height = 100
        screen_width = parent.winfo_screenwidth()
        screen_height = parent.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.top.geometry(f'{window_width}x{window_height}+{x}+{y}')
        
        # 항상 최상위에 표시
        self.top.transient(parent)
        self.top.grab_set()
        
        # 진행 상태 표시
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(
            self.top, 
            variable=self.progress_var,
            maximum=100,
            mode='determinate',
            length=200
        )
        self.progress.pack(pady=10)
        
        # 현재 작업 설명
        self.status_label = ttk.Label(self.top, text="파일 로딩 중...")
        self.status_label.pack(pady=5)
        
        # 진행률 텍스트
        self.percentage_label = ttk.Label(self.top, text="0%")
        self.percentage_label.pack(pady=5)
        
        # 창 닫기 버튼 비활성화
        self.top.protocol("WM_DELETE_WINDOW", lambda: None)
        
    def update_progress(self, value, status_text=None):
        self.progress_var.set(value)
        self.percentage_label.config(text=f"{int(value)}%")
        if status_text:
            self.status_label.config(text=status_text)
        self.top.update()
    
    def close(self):
        self.top.grab_release()
        self.top.destroy()

def main():
    app = DBManager()
    app.window.mainloop()

if __name__ == "__main__":
    main()